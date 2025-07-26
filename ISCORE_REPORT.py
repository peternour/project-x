import os
import shutil
import csv
from datetime import datetime
import openpyxl

# ================================
# USER CONFIGURATION SECTION
# ================================

# Source Excel file path
SOURCE_EXCEL = r"C:\Users\Peter\Downloads\Iscour_Report.xlsx"

# Destination directory
DEST_DIR = r"C:\Users\Peter\Downloads\New folder"

# Text file to sheet mapping
TEXT_FILE_MAPPING = {
    # Format: "text_file_path": ["sheet_name", ["col1", "col2", ...]]
    r"C:\Users\Peter\Downloads\عميل.txt": ["جديد", ["B", "C", "D", "E","F","G", "H"]],
    r"C:\Users\Peter\Downloads\ضامن.txt": ["الضامنين", ["B", "C", "D", "E", "F", "G"]]
}

# Starting row for data insertion (1-based index)
START_ROW = 7

# CSV delimiter (usually ',' for CSV)
DELIMITER = ','

# ================================
# END OF USER CONFIGURATION
# ================================

def copy_and_modify_workbook():
    # Create destination filename with timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d%H%M%S")
    dest_filename = f"{timestamp}_Iscore_Report.xlsx"
    dest_path = os.path.join(DEST_DIR, dest_filename)
    
    # Copy source workbook to destination
    shutil.copyfile(SOURCE_EXCEL, dest_path)
    print(f"Copied workbook to: {dest_path}")
    
    # Load the copied workbook
    workbook = openpyxl.load_workbook(dest_path)
    
    # Process each text file
    for text_file, (sheet_name, columns) in TEXT_FILE_MAPPING.items():
        if not os.path.exists(text_file):
            print(f"⚠️ Text file not found: {text_file}")
            continue
            
        if sheet_name not in workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found! Creating new sheet.")
            workbook.create_sheet(sheet_name)
        
        sheet = workbook[sheet_name]
        print(f"Processing {text_file} -> Sheet '{sheet_name}' Columns {columns}")
        
        # Read text file with UTF-8 encoding and write to specified columns
        try:
            with open(text_file, 'r', encoding='utf-8') as file:
                reader = csv.reader(file, delimiter=DELIMITER)
                for row_idx, row in enumerate(reader, start=START_ROW):
                    for col_idx, col_letter in enumerate(columns):
                        if col_idx < len(row):
                            cell = sheet[f"{col_letter}{row_idx}"]
                            # Preserve existing cell if no data in CSV
                            if row[col_idx].strip():
                                cell.value = row[col_idx]
        except UnicodeDecodeError:
            print(f"⚠️ Encoding error in {text_file}! Trying 'utf-8-sig'...")
            with open(text_file, 'r', encoding='utf-8-sig') as file:
                reader = csv.reader(file, delimiter=DELIMITER)
                for row_idx, row in enumerate(reader, start=START_ROW):
                    for col_idx, col_letter in enumerate(columns):
                        if col_idx < len(row):
                            cell = sheet[f"{col_letter}{row_idx}"]
                            if row[col_idx].strip():
                                cell.value = row[col_idx]
    
    # Save changes
    workbook.save(dest_path)
    print("Data transfer complete. Workbook saved with Arabic text support.")

if __name__ == "__main__":
    copy_and_modify_workbook()