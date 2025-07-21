import openpyxl
from openpyxl.utils import column_index_from_string
import os
import shutil
from datetime import datetime


def append_to_excel(text_file, excel_file, sheet_name, target_columns, text_encoding='utf-8'):
    """
    Appends text data to specific columns in an Excel sheet with duplicate checking
    
    Args:
        text_file (str): Path to input text file
        excel_file (str): Path to Excel file
        sheet_name (str): Worksheet name
        target_columns (list): Excel column letters
        text_encoding (str): Text file encoding
    """
    # Convert column letters to indices
    col_indices = [column_index_from_string(col) for col in target_columns]
    
    # Load or create workbook
    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
    else:
        wb = openpyxl.Workbook()
        if wb.sheetnames and sheet_name != wb.sheetnames[0]:
            wb.remove(wb.active)
    
    # Get or create worksheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    
    # Collect existing data for duplicate checking
    existing_combinations = set()
    if ws.max_row > 0:
        for row in range(1, ws.max_row + 1):
            values = []
            for col_idx in col_indices:
                cell_value = ws.cell(row=row, column=col_idx).value
                values.append(str(cell_value) if cell_value is not None else '')
            existing_combinations.add(tuple(values))
    
    # Prepare to process new data
    duplicates = []
    new_rows = []
    next_row = ws.max_row + 1
    
    # Read and process text data
    with open(text_file, 'r', encoding=text_encoding) as f:
        for line_num, line in enumerate(f, 1):
            line = line.strip()
            if not line:
                continue
                
            values = [v.strip() for v in line.split(',')]
            if len(values) < len(target_columns):
                values += [''] * (len(target_columns) - len(values))
            values = values[:len(target_columns)]
            
            # Check for duplicates
            value_tuple = tuple(values)
            if value_tuple in existing_combinations:
                duplicates.append(f"Line {line_num}: {line}")
                continue
                
            new_rows.append((next_row, values))
            existing_combinations.add(value_tuple)
            next_row += 1
    
    # Show duplicate warnings
    if duplicates:
        print("\n⚠️ DUPLICATE DATA WARNING ⚠️")
        print(f"Found {len(duplicates)} entries that already exist:")
        for dup in duplicates:
            print(f"  - {dup}")
        print("These entries will NOT be appended.")
        
        # Ask for confirmation
        response = input("\nContinue appending non-duplicate data? (y/n): ")
        if response.lower() != 'y':
            print("Operation canceled.")
            return
    
    # Write new data
    for row_idx, values in new_rows:
        for i, col_idx in enumerate(col_indices):
            ws.cell(row=row_idx, column=col_idx, value=values[i])
    
    # Save if there's new data
    if new_rows:
        wb.save(excel_file)
        print(f"\n✅ Appended {len(new_rows)} rows to columns {', '.join(target_columns)}")
    else:
        print("\nNo new data to append.")
        

def create_monthly_from_template(text_file, base_dir, template_path, sheet_name, target_columns, text_encoding='utf-8'):
    """
    Creates a new Excel file by copying a template, renames it with timestamp,
    and populates with text data starting from ROW 2
    
    Args:
        text_file (str): Path to input text file
        base_dir (str): Base directory for monthly folders
        template_path (str): Path to template Excel file
        sheet_name (str): Worksheet name
        target_columns (list): Excel column letters
        text_encoding (str): Text file encoding
    """
    # Get current time
    now = datetime.now()
    
    # Create monthly folder (format: MM)
    month_folder = now.strftime("%m")
    output_dir = os.path.join(base_dir, month_folder)
    os.makedirs(output_dir, exist_ok=True)
    
    # Create filename (format: yyyy-MM-dd_4-hh mm ss.xlsx)
    file_name = now.strftime("%Y-%m-%d_4-%H %M %S") + ".xlsx"
    excel_file = os.path.join(output_dir, file_name)
    
    # Copy template to new location
    shutil.copy(template_path, excel_file)
    print(f"📋 Copied template to: {excel_file}")
    
    # Open the copied workbook
    wb = openpyxl.load_workbook(excel_file)
    
    # Get or create worksheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    
    # Convert column letters to indices
    col_indices = [column_index_from_string(col) for col in target_columns]
    
    # Read and write data starting from ROW 2
    row_idx = 2  # Start from row 2 to preserve headers
    with open(text_file, 'r', encoding=text_encoding) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
                
            values = [v.strip() for v in line.split(',')]
            if len(values) < len(target_columns):
                values += [''] * (len(target_columns) - len(values))
            values = values[:len(target_columns)]
            
            # Write to specified columns
            for i, col_idx in enumerate(col_indices):
                ws.cell(row=row_idx, column=col_idx, value=values[i])
            
            row_idx += 1
    
    # Save new file
    wb.save(excel_file)
    print(f"💾 Saved populated file: {excel_file}")
    print(f"📝 Added {row_idx - 2} rows of data starting from row 2")
    return excel_file

if __name__ == "__main__":
    # ===== CONFIGURATION =====
    text_file = r"C:\Users\Peter\Desktop\INFORM.txt"
    text_encoding = 'utf-8'  # Use 'cp1256' for Arabic Windows
    
    # Configuration for existing Excel file
    existing_excel = r"D:\Committee\05 Neg-List\05-02 Guarantors\قائمة إدراج الكيانات الإرهابية والإرهابيين 23 يونيو 2025_GUR.xlsx"
    existing_sheet = "الإرهابيين"
    existing_columns = ["B", "D"]  # Columns in existing file
    
    # Configuration for new monthly files
    base_monthly_dir = r"D:\COMMITTEE FORMES\Iscour_List\2025"  # Base directory
    template_file = r"D:\COMMITTEE FORMES\Miza-Monthly\Co_py\2023-11-01_4-.xlsx"  # Template path
    new_sheet = "Sheet1"
    new_columns = ["D", "C"]  # Columns in new files
    # =========================
    
    # Append to existing Excel file
    if existing_excel:
        print(f"Processing existing file: {existing_excel}")
        append_to_excel(text_file, existing_excel, existing_sheet, existing_columns, text_encoding)
    
    # Create new monthly Excel file from template
    print("\nCreating new monthly file from template...")
    new_file = create_monthly_from_template(
        text_file, 
        base_monthly_dir, 
        template_file,
        new_sheet, 
        new_columns, 
        text_encoding
    )
    
    # Append to the new file (optional)
    # print("\nAppending to new file...")
    # append_to_excel(text_file, new_file, new_sheet, new_columns, text_encoding)
    
    print("\nOperation completed successfully!")